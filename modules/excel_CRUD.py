import pandas
from openpyxl import load_workbook
from modules.global_variables import configParser, current_dirname

"""
read_from_excel(str, str, int) -> DataFrame

Функция 'read_from_excel' имеет в качестве аргументов
 'excel_book' - строка, содержащая имя файла эксель из которого мы читаем данные,
 'excel_sheet' - строка, содержащая имя страницы эксель из которой мы читаем данные,
 'columns' - целое число, количество столбцов которые мы читаем из источника.
 """

def read_from_excel(excel_book, excel_sheet, columns=None):
    result = pandas.read_excel(excel_book, excel_sheet, usecols="A:G").sort_values(by="t_stamp")
    # Читаем датафрейм из экселя используя внутренний метод Pandas
    return result.iloc[:,:columns] if columns is not None else result

def write_excel_existing(sheet_, dataframe_):
    wb = load_workbook(current_dirname + configParser.get("IO_files", "input_excel"))
    if sheet_ in wb.sheetnames:
        wb.remove(wb[sheet_])
    ws1 = wb.create_sheet(sheet_, 0)
    df = dataframe_.values.tolist()
    # header = configParser.get("IO_files", "default_columns").split(",")
    # ws1.append(header)
    for i in range(len(df)):
        ws1.append(df[i])
    wb.save(current_dirname + configParser.get("IO_files", "input_excel"))